# SPDX-License-Identifier: Apache-2.0

"""Tests for Excel-focused office conversions."""

from __future__ import annotations

import importlib.util
import sys
from types import SimpleNamespace

import pytest

OPENPYXL_LIBRARY_AVAILABLE = importlib.util.find_spec("openpyxl") is not None
PANDAS_LIBRARY_AVAILABLE = importlib.util.find_spec("pandas") is not None


def test_xlsx_to_pdf_renames_libreoffice_output(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter()
    intermediate = tmp_path / "libreoffice" / "sheet.pdf"
    intermediate.parent.mkdir(parents=True)
    intermediate.write_bytes(b"%PDF")
    monkeypatch.setattr(
        converter,
        "_run_libreoffice_conversion",
        lambda input_path, output_format, output_dir: intermediate,
    )

    output = tmp_path / "final" / "report.pdf"
    output.parent.mkdir(parents=True, exist_ok=True)
    result = converter.xlsx_to_pdf(sample_office_files.xlsx, output)

    assert result == output
    assert output.exists()
    assert not intermediate.exists()


def test_xlsx_to_pdf_returns_same_path_when_no_rename_needed(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter()
    output = tmp_path / "report.pdf"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(b"%PDF")
    monkeypatch.setattr(
        converter,
        "_run_libreoffice_conversion",
        lambda input_path, output_format, output_dir: output,
    )
    assert converter.xlsx_to_pdf(sample_office_files.xlsx, output) == output


def test_xlsx_to_csv_uses_openpyxl_active_sheet(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter(openpyxl=True)

    class FakeSheet:
        def iter_rows(self, *, values_only):
            assert values_only is True
            return iter([(1, 2), ("a", None)])

    fake_workbook = SimpleNamespace(active=FakeSheet())
    monkeypatch.setitem(
        sys.modules,
        "openpyxl",
        SimpleNamespace(load_workbook=lambda path, data_only=True: fake_workbook),
    )

    output = tmp_path / "sheet.csv"
    result = converter.xlsx_to_csv(sample_office_files.xlsx, output)

    assert result == output
    assert output.read_text(encoding="utf-8").splitlines() == ["1,2", "a,"]


def test_xlsx_to_csv_uses_requested_sheet_name(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter(openpyxl=True)
    calls: list[str] = []

    class FakeSheet:
        def iter_rows(self, *, values_only):
            return iter([("revenue", 100)])

    class FakeWorkbook(dict):
        active = FakeSheet()

        def __getitem__(self, item):
            calls.append(item)
            return FakeSheet()

    monkeypatch.setitem(
        sys.modules,
        "openpyxl",
        SimpleNamespace(load_workbook=lambda path, data_only=True: FakeWorkbook()),
    )

    output = tmp_path / "named.csv"
    converter.xlsx_to_csv(sample_office_files.xlsx, output, sheet="Summary")

    assert calls == ["Summary"]
    assert output.read_text(encoding="utf-8").strip() == "revenue,100"


def test_xlsx_to_csv_falls_back_when_openpyxl_raises(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter(openpyxl=True)
    monkeypatch.setitem(
        sys.modules,
        "openpyxl",
        SimpleNamespace(
            load_workbook=lambda path, data_only=True: (_ for _ in ()).throw(
                ValueError("invalid workbook")
            )
        ),
    )
    fallback = tmp_path / "fallback" / "sheet.csv"
    fallback.parent.mkdir(parents=True)
    fallback.write_text("fallback,data", encoding="utf-8")
    monkeypatch.setattr(
        converter,
        "_run_libreoffice_conversion",
        lambda input_path, output_format, output_dir: fallback,
    )

    output = tmp_path / "out" / "sheet.csv"
    output.parent.mkdir(parents=True, exist_ok=True)
    assert converter.xlsx_to_csv(sample_office_files.xlsx, output) == output
    assert output.read_text(encoding="utf-8") == "fallback,data"


def test_xlsx_to_csv_falls_back_when_requested_sheet_missing(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter(openpyxl=True)

    class MissingWorkbook(dict):
        active = object()

        def __getitem__(self, item):
            raise KeyError(item)

    monkeypatch.setitem(
        sys.modules,
        "openpyxl",
        SimpleNamespace(load_workbook=lambda path, data_only=True: MissingWorkbook()),
    )

    fallback = tmp_path / "fallback.csv"
    fallback.write_text("used,fallback", encoding="utf-8")
    monkeypatch.setattr(
        converter,
        "_run_libreoffice_conversion",
        lambda input_path, output_format, output_dir: fallback,
    )

    output = tmp_path / "missing.csv"
    converter.xlsx_to_csv(sample_office_files.xlsx, output, sheet="DoesNotExist")
    assert output.read_text(encoding="utf-8") == "used,fallback"


def test_xlsx_to_ods_uses_libreoffice(
    make_converter, monkeypatch, sample_office_files, tmp_path
) -> None:
    converter = make_converter()
    intermediate = tmp_path / "intermediate.ods"
    intermediate.write_bytes(b"ods")
    monkeypatch.setattr(
        converter,
        "_run_libreoffice_conversion",
        lambda input_path, output_format, output_dir: intermediate,
    )

    output = tmp_path / "converted" / "sheet.ods"
    output.parent.mkdir(parents=True, exist_ok=True)
    result = converter.xlsx_to_ods(sample_office_files.xlsx, output)

    assert result == output
    assert output.exists()


@pytest.mark.parametrize(
    ("source_format", "target_format", "expected_name"),
    [
        ("xlsx", "pdf", "xlsx_to_pdf"),
        ("xlsx", "csv", "xlsx_to_csv"),
        ("xlsx", "ods", "xlsx_to_ods"),
    ],
)
def test_excel_converters_are_registered(
    make_converter, source_format: str, target_format: str, expected_name: str
) -> None:
    converter = make_converter()
    method = converter.get_converter(source_format, target_format)
    assert method is not None
    assert method.__name__ == expected_name


@pytest.mark.parametrize("fixture_name", ["xlsx", "xls", "nested_xlsx"])
def test_detect_format_identifies_excel_extensions(
    make_converter, office_modules, sample_office_files, fixture_name: str
) -> None:
    converter = make_converter()
    source = getattr(sample_office_files, fixture_name)
    detected = converter.detect_format(source)
    expected = (
        office_modules.converter.OfficeFormat.XLSX
        if source.suffix == ".xlsx"
        else office_modules.converter.OfficeFormat.XLS
    )
    assert detected == expected


@pytest.mark.skipif(not OPENPYXL_LIBRARY_AVAILABLE, reason="openpyxl not installed")
def test_xlsx_to_csv_with_real_openpyxl(make_converter, tmp_path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Value"])
    sheet.append(["Revenue", 42])
    source = tmp_path / "real.xlsx"
    workbook.save(source)

    converter = make_converter(openpyxl=True)
    output = tmp_path / "real.csv"
    result = converter.xlsx_to_csv(source, output)

    assert result == output
    assert output.read_text(encoding="utf-8").splitlines() == [
        "Name,Value",
        "Revenue,42",
    ]


@pytest.mark.skipif(not OPENPYXL_LIBRARY_AVAILABLE, reason="openpyxl not installed")
def test_excel_processor_extracts_structured_sheet_data(office_modules, tmp_path) -> None:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Name", "Value", "Choice"])
    sheet.append(["Revenue", 100, "Yes"])
    sheet.append(["Cost", 40, "No"])
    sheet["B4"] = "=SUM(B2:B3)"
    sheet.merge_cells("A6:B6")
    sheet["A6"] = "Merged title"
    sheet.freeze_panes = "A2"

    workbook.properties.title = "Quarterly workbook"
    workbook.properties.creator = "Joseph"

    workbook.defined_names.add(
        DefinedName("RevenueRange", attr_text="'Summary'!$B$2:$B$3")
    )

    validation = DataValidation(type="list", formula1='"Yes,No"')
    sheet.add_data_validation(validation)
    validation.add("C2:C3")

    sheet.conditional_formatting.add(
        "B2:B3",
        CellIsRule(operator="greaterThan", formula=["50"], stopIfTrue=True),
    )

    chart = BarChart()
    chart.title = "Performance"
    data = Reference(sheet, min_col=2, min_row=1, max_row=3)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "E2")

    source = tmp_path / "structured.xlsx"
    workbook.save(source)

    processor = office_modules.excel.ExcelProcessor()
    processor.load(source)

    assert processor.extract_sheets() == ["Summary"]
    formulas = processor.extract_formulas()
    assert formulas["Summary!B4"] == "=SUM(B2:B3)"

    charts = processor.extract_charts()
    assert len(charts) == 1
    assert charts[0].title == "Performance"

    metadata = processor.get_metadata()
    assert metadata.title == "Quarterly workbook"
    assert metadata.author == "Joseph"

    structured = processor.extract_data("Summary")
    assert structured["merged_cells"] == ["A6:B6"]
    assert structured["named_ranges"][0]["name"] == "RevenueRange"
    assert structured["data_validations"][0]["range"] == "C2:C3"
    assert structured["conditional_formatting"][0]["range"] == "B2:B3"

    cell_map = {
        cell["coordinate"]: cell
        for cell in structured["cells"]
    }
    assert cell_map["B2"]["data_type"] == "number"
    assert cell_map["B4"]["formula"] == "=SUM(B2:B3)"
    assert cell_map["A6"]["merged"] is True


@pytest.mark.skipif(not OPENPYXL_LIBRARY_AVAILABLE, reason="openpyxl not installed")
@pytest.mark.skipif(not PANDAS_LIBRARY_AVAILABLE, reason="pandas not installed")
def test_excel_processor_to_dataframe_fills_merged_cells(office_modules, tmp_path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Name", "Value"])
    sheet.append(["Revenue", 100])
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Total"

    source = tmp_path / "frame.xlsx"
    workbook.save(source)

    processor = office_modules.excel.ExcelProcessor()
    processor.load(source)
    frame = processor.to_dataframe("Sheet1")

    assert list(frame.columns) == ["Name", "Value"]
    assert frame.iloc[0].tolist() == ["Revenue", 100]
    assert frame.iloc[1].tolist() == ["Total", "Total"]


@pytest.mark.skipif(not OPENPYXL_LIBRARY_AVAILABLE, reason="openpyxl not installed")
def test_excel_processor_create_and_merge_workbooks(office_modules, tmp_path) -> None:
    import openpyxl

    processor = office_modules.excel.ExcelProcessor()
    workbook = processor.create_workbook(
        {
            "Created": {
                "rows": [["Name", "Value"], ["Revenue", 125]],
                "merged_cells": ["A3:B3"],
                "named_ranges": [{"name": "CreatedRange", "reference": "B2"}],
            }
        },
        metadata={"title": "Created workbook", "author": "Agentic Brain"},
    )
    assert workbook.sheetnames == ["Created"]
    assert workbook["Created"]["B2"].value == 125
    assert processor.get_metadata().title == "Created workbook"

    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "SheetA"
    ws1["A1"] = "First"
    wb1.save(first)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "SheetA"
    ws2["A1"] = "Second"
    wb2.save(second)

    merged = processor.merge_workbooks([first, second])
    assert merged.sheetnames == ["SheetA", "SheetA_2"]
    assert merged["SheetA"]["A1"].value == "First"
    assert merged["SheetA_2"]["A1"].value == "Second"
