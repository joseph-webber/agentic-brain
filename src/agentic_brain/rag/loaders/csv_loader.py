# SPDX-License-Identifier: Apache-2.0
# Copyright 2024-2026 Agentic Brain Contributors
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""CSV and Excel loaders for RAG pipelines."""

import csv
import logging
from io import BytesIO, StringIO
from pathlib import Path
from typing import Optional

from ..exceptions import LoaderError
from .base import BaseLoader, LoadedDocument

logger = logging.getLogger(__name__)

# Check for openpyxl
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for pandas
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class CSVLoader(BaseLoader):
    """Load and extract content from CSV files.

    Example:
        loader = CSVLoader()
        doc = loader.load_document("data.csv")
        docs = loader.load_folder("data/")
    """

    def __init__(
        self,
        base_path: str = ".",
        delimiter: str = ",",
        encoding: str = "utf-8",
        max_rows: int = 10000,
    ):
        self.base_path = Path(base_path)
        self.delimiter = delimiter
        self.encoding = encoding
        self.max_rows = max_rows

    @property
    def source_name(self) -> str:
        return "csv"

    def authenticate(self) -> bool:
        return True

    def _csv_to_text(self, content: str) -> str:
        """Convert CSV content to readable text."""
        try:
            reader = csv.reader(StringIO(content), delimiter=self.delimiter)
            rows = []

            for i, row in enumerate(reader):
                if i == 0:
                    rows.append(" | ".join(row))
                    rows.append("-" * 40)
                elif i < self.max_rows:
                    rows.append(" | ".join(str(cell) for cell in row))

            return "\n".join(rows)
        except Exception as e:
            logger.error(f"CSV parsing failed: {e}")
            return content

    def load_document(self, doc_id: str) -> Optional[LoadedDocument]:
        """Load a single CSV file."""
        path = Path(doc_id)
        if not path.is_absolute():
            path = self.base_path / path

        try:
            if not path.exists():
                raise LoaderError(
                    "File not found",
                    context={"path": str(path), "loader": self.source_name},
                )

            with open(path, encoding=self.encoding, errors="replace") as f:
                raw_content = f.read()
        except LoaderError:
            raise
        except FileNotFoundError as exc:
            logger.exception("CSV file not found")
            raise LoaderError(
                "File not found",
                context={"path": str(path), "loader": self.source_name},
            ) from exc
        except PermissionError as exc:
            logger.exception("Permission denied reading CSV")
            raise LoaderError(
                "Permission denied",
                context={"path": str(path), "loader": self.source_name},
            ) from exc
        except OSError as exc:
            logger.exception("I/O error reading CSV")
            raise LoaderError(
                "I/O error reading CSV",
                context={"path": str(path), "loader": self.source_name},
            ) from exc

        try:
            content = self._csv_to_text(raw_content)
        except Exception as exc:
            logger.exception("CSV parsing failed")
            raise LoaderError(
                "Corrupt CSV file",
                context={"path": str(path), "loader": self.source_name},
            ) from exc

        return LoadedDocument(
            content=content,
            source=self.source_name,
            source_id=str(path),
            filename=path.name,
            mime_type="text/csv",
            size_bytes=len(raw_content.encode()),
            metadata={"path": str(path), "delimiter": self.delimiter},
        )

    def load_folder(
        self, folder_path: str, recursive: bool = True
    ) -> list[LoadedDocument]:
        """Load all CSV files from a folder."""
        docs = []
        path = Path(folder_path)
        if not path.is_absolute():
            path = self.base_path / path

        if not path.exists():
            return docs

        pattern = "**/*.csv" if recursive else "*.csv"

        for csv_path in path.glob(pattern):
            try:
                doc = self.load_document(str(csv_path))
            except LoaderError as exc:
                logger.error("Failed to load %s: %s", csv_path, exc)
                continue

            if doc:
                docs.append(doc)

        logger.info(f"Loaded {len(docs)} CSV files from {folder_path}")
        return docs

    def search(self, query: str, max_results: int = 50) -> list[LoadedDocument]:
        return []


class ExcelLoader(BaseLoader):
    """Load and extract content from Excel files (.xlsx, .xls).

    Example:
        loader = ExcelLoader()
        doc = loader.load_document("report.xlsx")
        docs = loader.load_folder("reports/")
    """

    def __init__(
        self,
        base_path: str = ".",
        sheet_names: Optional[list[str]] = None,
        max_rows: int = 10000,
    ):
        self.base_path = Path(base_path)
        self.sheet_names = sheet_names
        self.max_rows = max_rows

    @property
    def source_name(self) -> str:
        return "excel"

    def authenticate(self) -> bool:
        return True

    def _excel_to_text(self, excel_bytes: bytes) -> str:
        """Convert Excel content to readable text."""
        if PANDAS_AVAILABLE:
            try:
                excel_file = BytesIO(excel_bytes)
                xls = pd.ExcelFile(excel_file)
                parts = []

                sheets = self.sheet_names or xls.sheet_names

                for sheet in sheets:
                    if sheet in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet, nrows=self.max_rows)
                        parts.append(f"## Sheet: {sheet}\n")
                        parts.append(df.to_string(index=False))
                        parts.append("\n")

                return "\n".join(parts)
            except Exception as e:
                logger.error(f"Pandas Excel parsing failed: {e}")

        if OPENPYXL_AVAILABLE:
            try:
                wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True)
                parts = []

                sheets = self.sheet_names or wb.sheetnames

                for sheet_name in sheets:
                    if sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        parts.append(f"## Sheet: {sheet_name}\n")

                        row_count = 0
                        for row in sheet.iter_rows(values_only=True):
                            if row_count >= self.max_rows:
                                break
                            row_text = " | ".join(str(cell or "") for cell in row)
                            if row_text.strip():
                                parts.append(row_text)
                            row_count += 1

                        parts.append("\n")

                return "\n".join(parts)
            except Exception as e:
                logger.error(f"openpyxl Excel parsing failed: {e}")

        return "[Excel content - install pandas or openpyxl]"

    def load_document(self, doc_id: str) -> Optional[LoadedDocument]:
        """Load a single Excel file."""
        path = Path(doc_id)
        if not path.is_absolute():
            path = self.base_path / path

        try:
            if not path.exists():
                raise LoaderError(
                    "File not found",
                    context={"path": str(path), "loader": self.source_name},
                )

            if path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
                logger.warning(f"Not an Excel file: {path}")
                return None

            with open(path, "rb") as f:
                excel_bytes = f.read()

            content = self._excel_to_text(excel_bytes)

            return LoadedDocument(
                content=content,
                source=self.source_name,
                source_id=str(path),
                filename=path.name,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                size_bytes=len(excel_bytes),
                metadata={"path": str(path)},
            )
        except LoaderError:
            raise
        except FileNotFoundError as exc:
            logger.exception("Excel file not found")
            raise LoaderError(
                "File not found",
                context={"path": str(path), "loader": self.source_name},
            ) from exc
        except PermissionError as exc:
            logger.exception("Permission denied reading Excel")
            raise LoaderError(
                "Permission denied",
                context={"path": str(path), "loader": self.source_name},
            ) from exc
        except OSError as exc:
            logger.exception("I/O error reading Excel")
            raise LoaderError(
                "I/O error reading Excel",
                context={"path": str(path), "loader": self.source_name},
            ) from exc
        except Exception as exc:
            logger.exception("Excel parsing failed")
            raise LoaderError(
                "Corrupt Excel file",
                context={"path": str(path), "loader": self.source_name},
            ) from exc

    def load_folder(
        self, folder_path: str, recursive: bool = True
    ) -> list[LoadedDocument]:
        """Load all Excel files from a folder."""
        docs = []
        path = Path(folder_path)
        if not path.is_absolute():
            path = self.base_path / path

        if not path.exists():
            return docs

        patterns = ["**/*.xlsx", "**/*.xls"] if recursive else ["*.xlsx", "*.xls"]

        for pattern in patterns:
            for excel_path in path.glob(pattern):
                try:
                    doc = self.load_document(str(excel_path))
                except LoaderError as exc:
                    logger.error("Failed to load %s: %s", excel_path, exc)
                    continue

                if doc:
                    docs.append(doc)

        logger.info(f"Loaded {len(docs)} Excel files from {folder_path}")
        return docs

    def search(self, query: str, max_results: int = 50) -> list[LoadedDocument]:
        return []


__all__ = ["CSVLoader", "ExcelLoader"]
