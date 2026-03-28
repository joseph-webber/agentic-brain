"""Excel workbook processor built on top of ``openpyxl``.

This module provides a structured Excel processor for modern XLSX-family
formats. It focuses on the operations most useful to downstream document and
RAG pipelines:

* list worksheets
* extract typed cell data
* extract formulas
* detect charts and images
* expose workbook metadata
* convert sheets to pandas DataFrames
* create new workbooks
* merge multiple workbooks into one

The implementation is intentionally defensive and keeps ``openpyxl`` optional
at import time. Runtime operations raise an explicit dependency error if the
library is unavailable.
"""

from __future__ import annotations

import logging
from copy import copy, deepcopy
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from .exceptions import DocumentModelError, UnsupportedOfficeFormatError
from .models import (
    Cell,
    Chart,
    Comment,
    DocumentContent,
    DocumentStyle,
    Image,
    Metadata,
    OfficeFormat,
    Worksheet,
)

logger = logging.getLogger(__name__)


class _OpenpyxlStub:
    """Fallback placeholder used when openpyxl is unavailable."""


try:  # pragma: no cover - import availability depends on environment
    import openpyxl
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell as OpenpyxlCell
    from openpyxl.cell.cell import MergedCell
    from openpyxl.chart._chart import ChartBase
    from openpyxl.comments import Comment as OpenpyxlComment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation as OpenpyxlDataValidation
    from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only without dependency
    openpyxl = None
    load_workbook = None
    OpenpyxlWorkbook = _OpenpyxlStub
    OpenpyxlWorksheet = _OpenpyxlStub
    OpenpyxlCell = _OpenpyxlStub
    MergedCell = _OpenpyxlStub
    ChartBase = _OpenpyxlStub
    OpenpyxlComment = _OpenpyxlStub
    Alignment = _OpenpyxlStub
    Font = _OpenpyxlStub
    PatternFill = _OpenpyxlStub
    DefinedName = _OpenpyxlStub
    OpenpyxlDataValidation = _OpenpyxlStub
    OPENPYXL_AVAILABLE = False

try:  # pragma: no cover - import availability depends on environment
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only without dependency
    pd = None
    PANDAS_AVAILABLE = False


class ExcelProcessorError(DocumentModelError):
    """Base exception for Excel processor failures."""


class ExcelDependencyError(ExcelProcessorError):
    """Raised when an optional Excel dependency is missing."""


class ExcelNotFoundError(ExcelProcessorError):
    """Raised when the source workbook does not exist."""


class ExcelCorruptedError(ExcelProcessorError):
    """Raised when a workbook cannot be loaded or parsed."""


class ExcelSheetNotFoundError(ExcelProcessorError):
    """Raised when a requested worksheet cannot be found."""


class ExcelProcessor:
    """Read, inspect, create, and merge XLSX workbooks with ``openpyxl``."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    def __init__(
        self,
        path: str | Path | None = None,
        *,
        read_only: bool = False,
        data_only: bool = False,
        keep_vba: bool = False,
        keep_links: bool = True,
    ) -> None:
        self.read_only = read_only
        self.data_only = data_only
        self.keep_vba = keep_vba
        self.keep_links = keep_links
        self._path: Path | None = Path(path) if path is not None else None
        self._workbook: OpenpyxlWorkbook | None = None
        self._formula_workbook: OpenpyxlWorkbook | None = None

    # ------------------------------------------------------------------
    # Public read API
    # ------------------------------------------------------------------
    def parse(self, path: str | Path | None = None) -> DocumentContent:
        """Load a workbook and return normalized document content."""

        self.load(path)
        workbook = self._require_workbook()
        worksheets = [self._worksheet_model(sheet) for sheet in workbook.worksheets]
        charts = self.extract_charts()
        images = self.extract_images()
        comments = [
            cell.comment
            for worksheet in worksheets
            for cell in worksheet.cells.values()
            if cell.comment is not None
        ]

        return DocumentContent(
            format=OfficeFormat.XLSX,
            worksheets=worksheets,
            charts=charts,
            images=images,
            comments=comments,
            metadata=self.get_metadata(),
            document_properties={
                "sheet_count": len(worksheets),
                "sheet_names": self.extract_sheets(),
                "named_ranges": self.get_named_ranges(),
                "has_macros": self.keep_vba,
            },
        )

    def load(self, path: str | Path | None = None) -> OpenpyxlWorkbook:
        """Load a workbook from disk and cache it on the processor."""

        self._ensure_openpyxl()
        if path is not None:
            self._path = Path(path)

        source = self._require_path()
        self._validate_path(source)

        try:
            workbook = load_workbook(
                source,
                read_only=self.read_only,
                data_only=self.data_only,
                keep_vba=self.keep_vba,
                keep_links=self.keep_links,
            )
        except FileNotFoundError as exc:
            raise ExcelNotFoundError(f"Workbook not found: {source}") from exc
        except Exception as exc:  # pragma: no cover - exercised with real files
            raise ExcelCorruptedError(f"Failed to load workbook: {source}: {exc}") from exc

        self._workbook = workbook
        self._formula_workbook = None
        return workbook

    def close(self) -> None:
        """Release cached workbook handles."""

        workbook = self._workbook
        if workbook is not None and hasattr(workbook, "close"):
            workbook.close()

        formula_workbook = self._formula_workbook
        if formula_workbook is not None and hasattr(formula_workbook, "close"):
            formula_workbook.close()

        self._workbook = None
        self._formula_workbook = None

    def extract_sheets(self) -> list[str]:
        """Return the workbook sheet names."""

        return list(self._require_workbook().sheetnames)

    def get_sheet_names(self) -> list[str]:
        """Backward-compatible alias for :meth:`extract_sheets`."""

        return self.extract_sheets()

    def extract_sheet(self, sheet: str | Worksheet | OpenpyxlWorksheet) -> Worksheet:
        """Return a single worksheet mapped into the shared worksheet model."""

        return self._worksheet_model(self._resolve_worksheet(sheet))

    def extract_all_sheets(self) -> list[Worksheet]:
        """Return all worksheets as shared worksheet models."""

        workbook = self._require_workbook()
        return [self._worksheet_model(sheet) for sheet in workbook.worksheets]

    def extract_data(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet,
        *,
        include_empty: bool = False,
    ) -> dict[str, Any]:
        """Extract rich typed cell data from a worksheet.

        The returned payload includes:

        * worksheet dimensions
        * merged cell ranges
        * named ranges scoped to the sheet
        * data validation rules
        * conditional formatting rules
        * per-cell typed metadata
        """

        worksheet = self._resolve_worksheet(sheet)
        merged_lookup = self._merged_lookup(worksheet)
        validation_map = self._validation_map(self.get_data_validations(worksheet))
        conditional_map = self._conditional_formatting_map(
            self.get_conditional_formatting(worksheet)
        )

        cells: list[dict[str, Any]] = []
        for row in worksheet.iter_rows():
            for raw_cell in row:
                payload = self._cell_payload(
                    worksheet=worksheet,
                    raw_cell=raw_cell,
                    merged_lookup=merged_lookup,
                    validation_map=validation_map,
                    conditional_map=conditional_map,
                )
                if payload is None:
                    continue
                if (
                    not include_empty
                    and payload["value"] is None
                    and payload["formula"] is None
                    and not payload["merged"]
                    and not payload["has_comment"]
                ):
                    continue
                cells.append(payload)

        return {
            "sheet": worksheet.title,
            "dimensions": {
                "min_row": worksheet.min_row,
                "max_row": worksheet.max_row,
                "min_column": worksheet.min_column,
                "max_column": worksheet.max_column,
            },
            "frozen_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
            "merged_cells": self.get_merged_cells(worksheet),
            "named_ranges": self.get_named_ranges(worksheet.title),
            "data_validations": self.get_data_validations(worksheet),
            "conditional_formatting": self.get_conditional_formatting(worksheet),
            "cells": cells,
        }

    def extract_formulas(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet | None = None,
    ) -> dict[str, str]:
        """Extract formulas from the workbook or a single sheet."""

        workbook = self._formula_source_workbook()
        worksheets = (
            workbook.worksheets
            if sheet is None
            else [self._resolve_worksheet_from_workbook(workbook, sheet)]
        )

        formulas: dict[str, str] = {}
        for worksheet in worksheets:
            for row in worksheet.iter_rows():
                for raw_cell in row:
                    if isinstance(raw_cell, MergedCell):
                        continue
                    formula = self._cell_formula(raw_cell)
                    if formula is not None:
                        formulas[f"{worksheet.title}!{raw_cell.coordinate}"] = formula
        return formulas

    def extract_charts(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet | None = None,
    ) -> list[Chart]:
        """Detect charts in the workbook or in a single sheet."""

        workbook = self._require_workbook()
        worksheets = (
            workbook.worksheets
            if sheet is None
            else [self._resolve_worksheet_from_workbook(workbook, sheet)]
        )

        charts: list[Chart] = []
        for worksheet in worksheets:
            for chart in getattr(worksheet, "_charts", []):
                try:
                    charts.append(self._chart_model(worksheet, chart))
                except Exception as exc:  # pragma: no cover - defensive
                    logger.warning("Failed to serialize chart on %s: %s", worksheet.title, exc)
        return charts

    def extract_images(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet | None = None,
    ) -> list[Image]:
        """Extract embedded worksheet images when available."""

        workbook = self._require_workbook()
        worksheets = (
            workbook.worksheets
            if sheet is None
            else [self._resolve_worksheet_from_workbook(workbook, sheet)]
        )

        images: list[Image] = []
        for worksheet in worksheets:
            for image in getattr(worksheet, "_images", []):
                try:
                    images.append(self._image_model(worksheet, image))
                except Exception as exc:  # pragma: no cover - defensive
                    logger.warning("Failed to serialize image on %s: %s", worksheet.title, exc)
        return images

    def get_metadata(self) -> Metadata:
        """Return workbook metadata and custom properties when available."""

        workbook = self._require_workbook()
        properties = workbook.properties
        keywords = self._split_keywords(getattr(properties, "keywords", None))

        return Metadata(
            title=getattr(properties, "title", None),
            subject=getattr(properties, "subject", None),
            author=getattr(properties, "creator", None),
            company=getattr(properties, "company", None),
            category=getattr(properties, "category", None),
            keywords=keywords,
            created_at=getattr(properties, "created", None),
            modified_at=getattr(properties, "modified", None),
            last_printed_at=getattr(properties, "lastPrinted", None),
            revision=self._string_or_none(getattr(properties, "revision", None)),
            custom_properties=self._custom_properties(workbook),
        )

    def extract_metadata(self) -> Metadata:
        """Backward-compatible alias for :meth:`get_metadata`."""

        return self.get_metadata()

    def to_dataframe(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet,
        *,
        header: bool = True,
        fill_merged: bool = True,
    ) -> Any:
        """Convert a worksheet into a pandas DataFrame."""

        if not PANDAS_AVAILABLE:
            raise ExcelDependencyError("pandas is required for DataFrame conversion")

        worksheet = self._resolve_worksheet(sheet)
        matrix = self._trim_matrix(self._sheet_matrix(worksheet, fill_merged=fill_merged))
        if not matrix:
            return pd.DataFrame()

        if header and self._looks_like_header(matrix[0]):
            columns = [self._header_name(value, index) for index, value in enumerate(matrix[0])]
            return pd.DataFrame(matrix[1:], columns=columns)
        return pd.DataFrame(matrix)

    # ------------------------------------------------------------------
    # Workbook creation and merge API
    # ------------------------------------------------------------------
    def create_workbook(
        self,
        sheets: Mapping[str, Any] | None = None,
        *,
        metadata: Metadata | Mapping[str, Any] | None = None,
        path: str | Path | None = None,
    ) -> OpenpyxlWorkbook:
        """Create a new workbook and optionally populate it."""

        self._ensure_openpyxl()
        workbook = openpyxl.Workbook()

        if sheets:
            default_sheet = workbook.active
            for index, (sheet_name, payload) in enumerate(sheets.items()):
                worksheet = default_sheet if index == 0 else workbook.create_sheet()
                worksheet.title = sheet_name
                self._populate_sheet(workbook, worksheet, payload)

        self._apply_metadata(workbook, metadata)
        self._workbook = workbook
        self._formula_workbook = None
        if path is not None:
            self.save(path)
        return workbook

    def merge_workbooks(
        self,
        paths: Sequence[str | Path],
        *,
        output_path: str | Path | None = None,
    ) -> OpenpyxlWorkbook:
        """Merge multiple XLSX-family workbooks into a single workbook."""

        self._ensure_openpyxl()
        if not paths:
            raise ValueError("At least one workbook path is required")

        merged = openpyxl.Workbook()
        merged.remove(merged.active)
        first_properties: Any | None = None

        for source in paths:
            source_path = Path(source)
            self._validate_path(source_path)
            workbook = load_workbook(
                source_path,
                read_only=False,
                data_only=False,
                keep_vba=False,
                keep_links=self.keep_links,
            )

            if first_properties is None:
                first_properties = copy(workbook.properties)

            sheet_name_map: dict[str, str] = {}
            for worksheet in workbook.worksheets:
                new_title = self._unique_sheet_name(worksheet.title, merged.sheetnames)
                sheet_name_map[worksheet.title] = new_title
                target = merged.create_sheet(new_title)
                self._copy_worksheet(worksheet, target)

            self._copy_named_ranges(
                source_workbook=workbook,
                target_workbook=merged,
                sheet_name_map=sheet_name_map,
                namespace=source_path.stem,
            )

            if hasattr(workbook, "close"):
                workbook.close()

        if first_properties is not None:
            merged.properties = first_properties

        self._workbook = merged
        self._formula_workbook = None
        if output_path is not None:
            self.save(output_path)
        return merged

    def save(self, path: str | Path) -> None:
        """Persist the current workbook to disk."""

        workbook = self._require_workbook()
        destination = Path(path)
        workbook.save(destination)
        self._path = destination

    # ------------------------------------------------------------------
    # Rich workbook feature accessors
    # ------------------------------------------------------------------
    def get_merged_cells(self, sheet: str | Worksheet | OpenpyxlWorksheet) -> list[str]:
        """Return merged ranges for a worksheet."""

        worksheet = self._resolve_worksheet(sheet)
        return [str(merged_range) for merged_range in worksheet.merged_cells.ranges]

    def get_named_ranges(self, sheet: str | None = None) -> list[dict[str, Any]]:
        """Return workbook defined names, optionally filtered to a worksheet."""

        workbook = self._require_workbook()
        ranges: list[dict[str, Any]] = []
        for defined_name in self._iter_defined_names(workbook):
            reference = self._defined_name_text(defined_name)
            if not reference:
                continue

            scope = self._defined_name_scope(workbook, defined_name)
            if sheet is not None and not self._defined_name_matches_sheet(
                reference, sheet, scope
            ):
                continue

            ranges.append(
                {
                    "name": getattr(defined_name, "name", None),
                    "reference": reference,
                    "scope": scope,
                    "comment": getattr(defined_name, "comment", None),
                    "hidden": bool(getattr(defined_name, "hidden", False)),
                }
            )
        return ranges

    def get_data_validations(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet,
    ) -> list[dict[str, Any]]:
        """Return sheet data validation rules."""

        worksheet = self._resolve_worksheet(sheet)
        validations: list[dict[str, Any]] = []
        data_validations = getattr(getattr(worksheet, "data_validations", None), "dataValidation", [])

        for validation in data_validations:
            validations.append(
                {
                    "range": str(validation.sqref),
                    "type": getattr(validation, "type", None),
                    "operator": getattr(validation, "operator", None),
                    "formula1": getattr(validation, "formula1", None),
                    "formula2": getattr(validation, "formula2", None),
                    "allow_blank": bool(getattr(validation, "allowBlank", False)),
                    "show_dropdown": not bool(getattr(validation, "showDropDown", False)),
                    "show_input_message": bool(
                        getattr(validation, "showInputMessage", False)
                    ),
                    "input_title": getattr(validation, "promptTitle", None),
                    "input_message": getattr(validation, "prompt", None),
                    "show_error_message": bool(
                        getattr(validation, "showErrorMessage", False)
                    ),
                    "error_title": getattr(validation, "errorTitle", None),
                    "error_message": getattr(validation, "error", None),
                }
            )
        return validations

    def get_conditional_formatting(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet,
    ) -> list[dict[str, Any]]:
        """Return conditional formatting rules for a worksheet."""

        worksheet = self._resolve_worksheet(sheet)
        rules: list[dict[str, Any]] = []
        raw_rules = getattr(worksheet.conditional_formatting, "_cf_rules", {})

        for range_object, entries in raw_rules.items():
            range_string = str(getattr(range_object, "sqref", range_object))
            for entry in entries:
                formulas = list(getattr(entry, "formula", []) or [])
                rules.append(
                    {
                        "range": range_string,
                        "type": getattr(entry, "type", None),
                        "operator": getattr(entry, "operator", None),
                        "formula": formulas,
                        "priority": getattr(entry, "priority", None),
                        "text": getattr(entry, "text", None),
                    }
                )
        return rules

    # ------------------------------------------------------------------
    # Internal conversion helpers
    # ------------------------------------------------------------------
    def _worksheet_model(self, worksheet: OpenpyxlWorksheet) -> Worksheet:
        """Convert an openpyxl worksheet into the shared worksheet model."""

        merged_lookup = self._merged_lookup(worksheet)
        validation_map = self._validation_map(self.get_data_validations(worksheet))

        cells: dict[str, Cell] = {}
        for row in worksheet.iter_rows():
            for raw_cell in row:
                cell_model = self._cell_model(
                    worksheet=worksheet,
                    raw_cell=raw_cell,
                    merged_lookup=merged_lookup,
                    validation_map=validation_map,
                )
                if cell_model is None:
                    continue
                if (
                    cell_model.value is None
                    and cell_model.formula is None
                    and not cell_model.merged
                    and cell_model.comment is None
                    and not cell_model.data_validation
                ):
                    continue
                cells[cell_model.reference] = cell_model

        return Worksheet(
            name=worksheet.title,
            cells=cells,
            column_widths=self._column_widths(worksheet),
            row_heights=self._row_heights(worksheet),
            frozen_panes=str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
            tab_color=self._color_to_hex(getattr(worksheet.sheet_properties.tabColor, "rgb", None))
            if getattr(worksheet.sheet_properties, "tabColor", None)
            else None,
            charts=[self._chart_model(worksheet, chart) for chart in getattr(worksheet, "_charts", [])],
            protection=self._sheet_protection(worksheet),
        )

    def _cell_model(
        self,
        *,
        worksheet: OpenpyxlWorksheet,
        raw_cell: Any,
        merged_lookup: dict[str, dict[str, Any]],
        validation_map: dict[str, list[dict[str, Any]]],
    ) -> Cell | None:
        coordinate = getattr(raw_cell, "coordinate", None)
        if coordinate is None:
            return None

        merge_info = merged_lookup.get(coordinate, {})
        anchor_coordinate = merge_info.get("anchor", coordinate)
        anchor_cell = worksheet[anchor_coordinate]
        formula = self._cell_formula(anchor_cell) if merge_info.get("is_anchor", True) else None
        comment = self._comment_model(anchor_cell.comment) if anchor_cell.comment else None
        validations = validation_map.get(coordinate, [])

        return Cell(
            reference=coordinate,
            value=self._normalize_value(None if isinstance(raw_cell, MergedCell) else raw_cell.value),
            formula=formula,
            style=self._style_from_cell(anchor_cell),
            comment=comment,
            data_validation=validations[0] if validations else {},
            hyperlink=getattr(getattr(anchor_cell, "hyperlink", None), "target", None)
            or getattr(anchor_cell, "hyperlink", None),
            merged=bool(merge_info),
            merge_range=merge_info.get("range"),
            number_format=getattr(anchor_cell, "number_format", None),
        )

    def _cell_payload(
        self,
        *,
        worksheet: OpenpyxlWorksheet,
        raw_cell: Any,
        merged_lookup: dict[str, dict[str, Any]],
        validation_map: dict[str, list[dict[str, Any]]],
        conditional_map: dict[str, list[dict[str, Any]]],
    ) -> dict[str, Any] | None:
        coordinate = getattr(raw_cell, "coordinate", None)
        if coordinate is None:
            return None

        merge_info = merged_lookup.get(coordinate, {})
        anchor_coordinate = merge_info.get("anchor", coordinate)
        anchor_cell = worksheet[anchor_coordinate]
        value = None if isinstance(raw_cell, MergedCell) else self._normalize_value(raw_cell.value)
        formula = self._cell_formula(anchor_cell) if merge_info.get("is_anchor", True) else None
        comment = self._comment_model(anchor_cell.comment) if anchor_cell.comment else None

        return {
            "coordinate": coordinate,
            "row": getattr(raw_cell, "row", anchor_cell.row),
            "column": getattr(raw_cell, "column", anchor_cell.column),
            "value": value,
            "display_value": "" if value is None else str(value),
            "data_type": self._cell_type(anchor_cell, raw_cell),
            "python_type": type(value).__name__ if value is not None else None,
            "formula": formula,
            "number_format": getattr(anchor_cell, "number_format", None),
            "hyperlink": getattr(getattr(anchor_cell, "hyperlink", None), "target", None)
            or getattr(anchor_cell, "hyperlink", None),
            "has_comment": comment is not None,
            "comment": {"author": comment.author, "text": comment.text} if comment else None,
            "merged": bool(merge_info),
            "merge_range": merge_info.get("range"),
            "merge_anchor": anchor_coordinate if merge_info else None,
            "data_validation": validation_map.get(coordinate, []),
            "conditional_formatting": conditional_map.get(coordinate, []),
        }

    def _chart_model(self, worksheet: OpenpyxlWorksheet, chart: Any) -> Chart:
        """Map an openpyxl chart object into the shared chart model."""

        return Chart(
            chart_type=type(chart).__name__.replace("Chart", "").lower() or "chart",
            title=self._chart_title(chart),
            series=self._chart_series(chart),
            legend={"position": self._string_or_none(getattr(getattr(chart, "legend", None), "position", None))},
            position={
                "sheet": worksheet.title,
                "anchor": self._chart_anchor(chart),
            },
            axis_titles={
                "x": self._chart_axis_title(getattr(chart, "x_axis", None)),
                "y": self._chart_axis_title(getattr(chart, "y_axis", None)),
            },
            data_range=self._chart_data_range(chart),
        )

    def _image_model(self, worksheet: OpenpyxlWorksheet, image: Any) -> Image:
        """Map an openpyxl image into the shared image model."""

        mime_type = "image/png"
        fmt = self._string_or_none(getattr(image, "format", None))
        if fmt:
            mime_type = f"image/{fmt.lower()}"

        anchor = self._anchor_coordinate(getattr(image, "anchor", None))
        data = image._data() if hasattr(image, "_data") else b""
        return Image(
            data=data,
            mime_type=mime_type,
            width=float(getattr(image, "width", 0) or 0) or None,
            height=float(getattr(image, "height", 0) or 0) or None,
            position={"sheet": worksheet.title, "anchor": anchor} if anchor else {"sheet": worksheet.title},
        )

    # ------------------------------------------------------------------
    # Workbook write helpers
    # ------------------------------------------------------------------
    def _populate_sheet(
        self,
        workbook: OpenpyxlWorkbook,
        worksheet: OpenpyxlWorksheet,
        payload: Any,
    ) -> None:
        if isinstance(payload, Worksheet):
            self._write_worksheet_model(worksheet, payload)
            return

        if isinstance(payload, Mapping):
            rows = payload.get("rows", [])
            for row in rows:
                worksheet.append(list(row))

            for merged_range in payload.get("merged_cells", []):
                worksheet.merge_cells(str(merged_range))

            freeze_panes = payload.get("freeze_panes")
            if freeze_panes:
                worksheet.freeze_panes = freeze_panes

            for validation in payload.get("data_validations", []):
                self._add_validation_from_mapping(worksheet, validation)

            for named_range in payload.get("named_ranges", []):
                reference = named_range.get("reference")
                name = named_range.get("name")
                if reference and name:
                    self._add_defined_name(
                        workbook,
                        DefinedName(
                            name=name,
                            attr_text=self._qualify_reference(worksheet.title, reference),
                        ),
                    )
            return

        for row in payload or []:
            worksheet.append(list(row))

    def _write_worksheet_model(
        self,
        worksheet: OpenpyxlWorksheet,
        model: Worksheet,
    ) -> None:
        for cell in sorted(model.cells.values(), key=lambda item: self._coordinate_sort_key(item.reference)):
            target = worksheet[cell.reference]
            target.value = cell.formula if cell.formula is not None else cell.value
            target.number_format = cell.number_format or target.number_format
            if cell.hyperlink:
                target.hyperlink = cell.hyperlink
            if cell.comment is not None:
                target.comment = OpenpyxlComment(cell.comment.text, cell.comment.author)
            self._apply_style_to_cell(target, cell.style)

        for merge_range in sorted(
            {
                cell.merge_range
                for cell in model.cells.values()
                if cell.merged and cell.merge_range
            }
        ):
            worksheet.merge_cells(merge_range)

        for index, width in model.column_widths.items():
            worksheet.column_dimensions[get_column_letter(index)].width = width
        for index, height in model.row_heights.items():
            worksheet.row_dimensions[index].height = height
        if model.frozen_panes:
            worksheet.freeze_panes = model.frozen_panes

    def _add_validation_from_mapping(
        self,
        worksheet: OpenpyxlWorksheet,
        payload: Mapping[str, Any],
    ) -> None:
        validation = OpenpyxlDataValidation(
            type=payload.get("type"),
            operator=payload.get("operator"),
            formula1=payload.get("formula1"),
            formula2=payload.get("formula2"),
            allow_blank=bool(payload.get("allow_blank", False)),
        )
        validation.promptTitle = payload.get("input_title")
        validation.prompt = payload.get("input_message")
        validation.errorTitle = payload.get("error_title")
        validation.error = payload.get("error_message")
        validation.showInputMessage = bool(payload.get("show_input_message", False))
        validation.showErrorMessage = bool(payload.get("show_error_message", False))
        validation.showDropDown = not bool(payload.get("show_dropdown", True))
        worksheet.add_data_validation(validation)
        validation.add(str(payload.get("range")))

    def _apply_metadata(
        self,
        workbook: OpenpyxlWorkbook,
        metadata: Metadata | Mapping[str, Any] | None,
    ) -> None:
        if metadata is None:
            return

        if isinstance(metadata, Mapping):
            payload = dict(metadata)
        else:
            payload = {
                "title": metadata.title,
                "subject": metadata.subject,
                "author": metadata.author,
                "category": metadata.category,
                "keywords": metadata.keywords,
                "created_at": metadata.created_at,
                "modified_at": metadata.modified_at,
                "last_printed_at": metadata.last_printed_at,
                "revision": metadata.revision,
            }
        properties = workbook.properties
        properties.title = payload.get("title")
        properties.subject = payload.get("subject")
        properties.creator = payload.get("author")
        properties.category = payload.get("category")
        properties.keywords = ",".join(payload.get("keywords", []) or [])
        properties.created = payload.get("created_at")
        properties.modified = payload.get("modified_at")
        properties.lastPrinted = payload.get("last_printed_at")
        revision = payload.get("revision")
        properties.revision = None if revision is None else str(revision)

    def _copy_worksheet(
        self,
        source: OpenpyxlWorksheet,
        target: OpenpyxlWorksheet,
    ) -> None:
        for row in source.iter_rows():
            for source_cell in row:
                if isinstance(source_cell, MergedCell):
                    continue

                target_cell = target[source_cell.coordinate]
                target_cell.value = source_cell.value
                if getattr(source_cell, "has_style", False):
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

        for key, dimension in source.column_dimensions.items():
            target_dimension = target.column_dimensions[key]
            target_dimension.width = dimension.width
            target_dimension.hidden = dimension.hidden
            target_dimension.bestFit = dimension.bestFit

        for key, dimension in source.row_dimensions.items():
            target_dimension = target.row_dimensions[key]
            target_dimension.height = dimension.height
            target_dimension.hidden = dimension.hidden

        if source.freeze_panes:
            target.freeze_panes = source.freeze_panes
        target.protection = copy(source.protection)
        if getattr(source.sheet_properties, "tabColor", None) is not None:
            target.sheet_properties.tabColor = copy(source.sheet_properties.tabColor)
        if getattr(source.auto_filter, "ref", None):
            target.auto_filter.ref = source.auto_filter.ref

        for merged_range in source.merged_cells.ranges:
            target.merge_cells(str(merged_range))

        for validation in getattr(getattr(source, "data_validations", None), "dataValidation", []):
            target.add_data_validation(deepcopy(validation))

        raw_rules = getattr(source.conditional_formatting, "_cf_rules", {})
        for range_object, entries in raw_rules.items():
            range_string = str(getattr(range_object, "sqref", range_object))
            for entry in entries:
                target.conditional_formatting.add(range_string, deepcopy(entry))

        for table in getattr(source, "tables", {}).values():
            cloned_table = deepcopy(table)
            new_name = self._unique_defined_name(
                getattr(cloned_table, "displayName", "Table"),
                {name.lower() for name in getattr(target, "tables", {}).keys()},
            )
            cloned_table.name = new_name
            cloned_table.displayName = new_name
            target.add_table(cloned_table)

        for chart in getattr(source, "_charts", []):
            try:
                target.add_chart(deepcopy(chart), self._chart_anchor(chart) or "A1")
            except Exception as exc:  # pragma: no cover - defensive
                logger.debug("Skipping chart copy on %s: %s", source.title, exc)

        for image in getattr(source, "_images", []):
            try:
                target.add_image(deepcopy(image), self._anchor_coordinate(getattr(image, "anchor", None)) or "A1")
            except Exception as exc:  # pragma: no cover - defensive
                logger.debug("Skipping image copy on %s: %s", source.title, exc)

    def _copy_named_ranges(
        self,
        *,
        source_workbook: OpenpyxlWorkbook,
        target_workbook: OpenpyxlWorkbook,
        sheet_name_map: Mapping[str, str],
        namespace: str,
    ) -> None:
        existing = {name.lower() for name in getattr(target_workbook, "defined_names", {}).keys()}

        for defined_name in self._iter_defined_names(source_workbook):
            reference = self._defined_name_text(defined_name)
            if not reference:
                continue

            rewritten_reference = self._rewrite_defined_name_reference(reference, sheet_name_map)
            name = getattr(defined_name, "name", "Name")
            if name.lower() in existing:
                name = self._unique_defined_name(f"{namespace}_{name}", existing)

            local_sheet_id = getattr(defined_name, "localSheetId", None)
            if local_sheet_id is not None:
                try:
                    source_sheet_name = source_workbook.sheetnames[int(local_sheet_id)]
                    target_sheet_name = sheet_name_map.get(source_sheet_name, source_sheet_name)
                    local_sheet_id = target_workbook.sheetnames.index(target_sheet_name)
                except (IndexError, TypeError, ValueError):
                    local_sheet_id = None

            cloned = DefinedName(
                name=name,
                attr_text=rewritten_reference,
                comment=getattr(defined_name, "comment", None),
                hidden=getattr(defined_name, "hidden", False),
                localSheetId=local_sheet_id,
            )
            self._add_defined_name(target_workbook, cloned)
            existing.add(name.lower())

    # ------------------------------------------------------------------
    # Internal utility helpers
    # ------------------------------------------------------------------
    def _ensure_openpyxl(self) -> None:
        if not OPENPYXL_AVAILABLE:
            raise ExcelDependencyError(
                "openpyxl is required for Excel processing. Install agentic-brain[documents]"
            )

    def _require_workbook(self) -> OpenpyxlWorkbook:
        if self._workbook is None:
            raise ExcelProcessorError("No workbook loaded. Call load(), parse(), or create_workbook() first.")
        return self._workbook

    def _require_path(self) -> Path:
        if self._path is None:
            raise ExcelProcessorError("No workbook path configured.")
        return self._path

    def _validate_path(self, path: Path) -> None:
        if not path.exists():
            raise ExcelNotFoundError(f"Workbook not found: {path}")
        if not path.is_file():
            raise ExcelNotFoundError(f"Workbook path is not a file: {path}")
        if path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise UnsupportedOfficeFormatError(path.suffix.lower())

    def _resolve_worksheet(
        self,
        sheet: str | Worksheet | OpenpyxlWorksheet,
    ) -> OpenpyxlWorksheet:
        workbook = self._require_workbook()
        return self._resolve_worksheet_from_workbook(workbook, sheet)

    def _resolve_worksheet_from_workbook(
        self,
        workbook: OpenpyxlWorkbook,
        sheet: str | Worksheet | OpenpyxlWorksheet,
    ) -> OpenpyxlWorksheet:
        if isinstance(sheet, str):
            if sheet not in workbook.sheetnames:
                raise ExcelSheetNotFoundError(f"Worksheet not found: {sheet}")
            return workbook[sheet]
        if isinstance(sheet, Worksheet):
            if sheet.name not in workbook.sheetnames:
                raise ExcelSheetNotFoundError(f"Worksheet not found: {sheet.name}")
            return workbook[sheet.name]
        return sheet

    def _formula_source_workbook(self) -> OpenpyxlWorkbook:
        if not self.data_only:
            return self._require_workbook()

        if self._formula_workbook is None:
            self._ensure_openpyxl()
            source = self._require_path()
            self._formula_workbook = load_workbook(
                source,
                read_only=self.read_only,
                data_only=False,
                keep_vba=self.keep_vba,
                keep_links=self.keep_links,
            )
        return self._formula_workbook

    def _column_widths(self, worksheet: OpenpyxlWorksheet) -> dict[int, float]:
        widths: dict[int, float] = {}
        for key, dimension in worksheet.column_dimensions.items():
            width = getattr(dimension, "width", None)
            if width is not None:
                widths[openpyxl.utils.column_index_from_string(key)] = float(width)
        return widths

    def _row_heights(self, worksheet: OpenpyxlWorksheet) -> dict[int, float]:
        heights: dict[int, float] = {}
        for key, dimension in worksheet.row_dimensions.items():
            height = getattr(dimension, "height", None)
            if height is not None:
                heights[int(key)] = float(height)
        return heights

    def _sheet_protection(self, worksheet: OpenpyxlWorksheet) -> dict[str, Any]:
        protection = getattr(worksheet, "protection", None)
        if protection is None:
            return {}
        return {
            "sheet": bool(getattr(protection, "sheet", False)),
            "objects": bool(getattr(protection, "objects", False)),
            "scenarios": bool(getattr(protection, "scenarios", False)),
            "format_cells": bool(getattr(protection, "formatCells", False)),
            "format_columns": bool(getattr(protection, "formatColumns", False)),
            "format_rows": bool(getattr(protection, "formatRows", False)),
        }

    def _style_from_cell(self, cell: OpenpyxlCell) -> DocumentStyle:
        font = getattr(cell, "font", None)
        fill = getattr(cell, "fill", None)
        alignment = getattr(cell, "alignment", None)
        return DocumentStyle(
            font_family=getattr(font, "name", None) or "Calibri",
            font_size=float(getattr(font, "size", 12.0) or 12.0),
            bold=bool(getattr(font, "bold", False)),
            italic=bool(getattr(font, "italic", False)),
            underline=bool(getattr(font, "underline", False)),
            text_color=self._color_to_hex(getattr(getattr(font, "color", None), "rgb", None))
            or "#000000",
            background_color=self._color_to_hex(
                getattr(getattr(fill, "fgColor", None), "rgb", None)
            )
            or "#FFFFFF",
            alignment=getattr(alignment, "horizontal", None) or "left",
            styles={
                "vertical_alignment": getattr(alignment, "vertical", None),
                "wrap_text": bool(getattr(alignment, "wrap_text", False)),
                "shrink_to_fit": bool(getattr(alignment, "shrink_to_fit", False)),
                "number_format": getattr(cell, "number_format", None),
            },
        )

    def _apply_style_to_cell(self, target: OpenpyxlCell, style: DocumentStyle) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        target.font = Font(
            name=style.font_family,
            size=style.font_size,
            bold=style.bold,
            italic=style.italic,
            underline="single" if style.underline else None,
            color=(style.text_color or "#000000").lstrip("#"),
        )
        target.alignment = Alignment(horizontal=style.alignment or "left")
        background = (style.background_color or "#FFFFFF").lstrip("#")
        target.fill = PatternFill(
            fill_type="solid",
            fgColor=background,
            bgColor=background,
        )

    def _comment_model(self, comment: Any) -> Comment:
        return Comment(author=getattr(comment, "author", ""), text=getattr(comment, "text", ""))

    def _cell_formula(self, cell: Any) -> str | None:
        value = getattr(cell, "value", None)
        if isinstance(value, str) and value.startswith("="):
            return value
        if getattr(cell, "data_type", None) == "f" and value is not None:
            value_string = str(value)
            return value_string if value_string.startswith("=") else f"={value_string}"
        return None

    def _cell_type(self, anchor_cell: Any, raw_cell: Any) -> str:
        if isinstance(raw_cell, MergedCell):
            return "merged"
        if getattr(anchor_cell, "is_date", False):
            return "date"
        mapping = {
            "b": "boolean",
            "d": "date",
            "e": "error",
            "f": "formula",
            "inlineStr": "string",
            "n": "number",
            "null": "empty",
            "s": "string",
            "str": "string",
        }
        return mapping.get(getattr(anchor_cell, "data_type", None), "unknown")

    def _normalize_value(self, value: Any) -> Any:
        if value == "":
            return ""
        return value

    def _merged_lookup(self, worksheet: OpenpyxlWorksheet) -> dict[str, dict[str, Any]]:
        lookup: dict[str, dict[str, Any]] = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            anchor = f"{get_column_letter(min_col)}{min_row}"
            for row_index in range(min_row, max_row + 1):
                for column_index in range(min_col, max_col + 1):
                    coordinate = f"{get_column_letter(column_index)}{row_index}"
                    lookup[coordinate] = {
                        "range": str(merged_range),
                        "anchor": anchor,
                        "is_anchor": coordinate == anchor,
                    }
        return lookup

    def _validation_map(
        self,
        validations: Iterable[dict[str, Any]],
    ) -> dict[str, list[dict[str, Any]]]:
        mapping: dict[str, list[dict[str, Any]]] = {}
        for validation in validations:
            for coordinate in self._expand_sqref(validation.get("range")):
                mapping.setdefault(coordinate, []).append(validation)
        return mapping

    def _conditional_formatting_map(
        self,
        rules: Iterable[dict[str, Any]],
    ) -> dict[str, list[dict[str, Any]]]:
        mapping: dict[str, list[dict[str, Any]]] = {}
        for rule in rules:
            for coordinate in self._expand_sqref(rule.get("range")):
                mapping.setdefault(coordinate, []).append(rule)
        return mapping

    def _expand_sqref(self, sqref: Any) -> list[str]:
        coordinates: list[str] = []
        if not sqref:
            return coordinates

        for token in str(sqref).split():
            if ":" not in token:
                coordinates.append(token)
                continue
            min_col, min_row, max_col, max_row = range_boundaries(token)
            for row_index in range(min_row, max_row + 1):
                for column_index in range(min_col, max_col + 1):
                    coordinates.append(f"{get_column_letter(column_index)}{row_index}")
        return coordinates

    def _sheet_matrix(
        self,
        worksheet: OpenpyxlWorksheet,
        *,
        fill_merged: bool,
    ) -> list[list[Any]]:
        merged_lookup = self._merged_lookup(worksheet)
        matrix: list[list[Any]] = []
        for row in worksheet.iter_rows():
            rendered_row: list[Any] = []
            for raw_cell in row:
                if isinstance(raw_cell, MergedCell) and fill_merged:
                    anchor = merged_lookup.get(raw_cell.coordinate, {}).get("anchor")
                    rendered_row.append(worksheet[anchor].value if anchor else None)
                else:
                    rendered_row.append(raw_cell.value)
            matrix.append(rendered_row)
        return matrix

    def _trim_matrix(self, matrix: list[list[Any]]) -> list[list[Any]]:
        trimmed = [list(row) for row in matrix]
        while trimmed and not any(value is not None and value != "" for value in trimmed[-1]):
            trimmed.pop()
        if not trimmed:
            return []

        max_width = 0
        for row in trimmed:
            for index in range(len(row) - 1, -1, -1):
                if row[index] is not None and row[index] != "":
                    max_width = max(max_width, index + 1)
                    break
        return [row[:max_width] for row in trimmed]

    def _looks_like_header(self, row: Sequence[Any]) -> bool:
        values = [value for value in row if value not in (None, "")]
        if not values:
            return False
        strings = [value for value in values if isinstance(value, str)]
        if len(strings) != len(values):
            return False
        normalized = [value.strip().lower() for value in strings if value.strip()]
        return len(normalized) == len(set(normalized))

    def _header_name(self, value: Any, index: int) -> str:
        if value is None or str(value).strip() == "":
            return f"column_{index + 1}"
        return str(value)

    def _chart_title(self, chart: Any) -> str | None:
        title = getattr(chart, "title", None)
        if title is None:
            return None
        if isinstance(title, str):
            return title

        text = getattr(getattr(title, "tx", None), "rich", None)
        if text is not None:
            fragments: list[str] = []
            for paragraph in getattr(text, "p", []):
                for run in getattr(paragraph, "r", []):
                    token = getattr(run, "t", None)
                    if token:
                        fragments.append(token)
            if fragments:
                return "".join(fragments)
        return self._string_or_none(title)

    def _chart_series(self, chart: Any) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []
        for index, series in enumerate(getattr(chart, "ser", []) or []):
            values_ref = self._series_reference(getattr(series, "val", None))
            categories_ref = self._series_reference(getattr(series, "cat", None))
            entries.append(
                {
                    "name": self._series_name(series) or f"Series {index + 1}",
                    "values_ref": values_ref,
                    "categories_ref": categories_ref,
                }
            )
        return entries

    def _series_name(self, series: Any) -> str | None:
        series_title = getattr(series, "title", None)
        if isinstance(series_title, str):
            return series_title

        tx = getattr(series, "tx", None)
        if tx is not None:
            string_ref = getattr(tx, "strRef", None)
            if string_ref is not None and getattr(string_ref, "f", None):
                return str(string_ref.f)
            if getattr(tx, "v", None):
                return str(tx.v)
        return None

    def _series_reference(self, source: Any) -> str | None:
        if source is None:
            return None
        for attribute in ("numRef", "strRef", "multiLvlStrRef"):
            ref = getattr(source, attribute, None)
            formula = getattr(ref, "f", None)
            if formula:
                return str(formula)
        formula = getattr(source, "f", None)
        if formula:
            return str(formula)
        return None

    def _chart_data_range(self, chart: Any) -> str | None:
        references = [
            reference
            for series in self._chart_series(chart)
            for reference in (series.get("values_ref"), series.get("categories_ref"))
            if reference
        ]
        return ", ".join(dict.fromkeys(references)) if references else None

    def _chart_axis_title(self, axis: Any) -> str | None:
        if axis is None:
            return None
        title = getattr(axis, "title", None)
        if title is None:
            return None
        return self._chart_title(type("AxisProxy", (), {"title": title})())

    def _chart_anchor(self, chart: Any) -> str | None:
        return self._anchor_coordinate(getattr(chart, "anchor", None))

    def _anchor_coordinate(self, anchor: Any) -> str | None:
        if anchor is None:
            return None
        if isinstance(anchor, str):
            return anchor
        source = getattr(anchor, "_from", None)
        if source is not None:
            return f"{get_column_letter(source.col + 1)}{source.row + 1}"
        marker = getattr(anchor, "from_", None)
        if marker is not None:
            return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
        return None

    def _iter_defined_names(self, workbook: OpenpyxlWorkbook) -> Iterable[Any]:
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return []
        if hasattr(defined_names, "values"):
            return list(defined_names.values())
        return list(getattr(defined_names, "definedName", []))

    def _defined_name_text(self, defined_name: Any) -> str | None:
        for attribute in ("attr_text", "value", "text"):
            value = getattr(defined_name, attribute, None)
            if value:
                return str(value)
        return None

    def _defined_name_scope(self, workbook: OpenpyxlWorkbook, defined_name: Any) -> str:
        local_sheet_id = getattr(defined_name, "localSheetId", None)
        if local_sheet_id is None:
            return "Workbook"
        try:
            return workbook.sheetnames[int(local_sheet_id)]
        except (IndexError, TypeError, ValueError):
            return "Workbook"

    def _defined_name_matches_sheet(self, reference: str, sheet: str, scope: str) -> bool:
        if scope not in {"Workbook", sheet}:
            return False
        return f"'{sheet}'!" in reference or f"{sheet}!" in reference or scope == sheet

    def _rewrite_defined_name_reference(
        self,
        reference: str,
        sheet_name_map: Mapping[str, str],
    ) -> str:
        rewritten = reference
        for source_name, target_name in sheet_name_map.items():
            rewritten = rewritten.replace(f"'{source_name}'!", f"'{target_name}'!")
            rewritten = rewritten.replace(f"{source_name}!", f"{target_name}!")
        return rewritten

    def _qualify_reference(self, sheet_name: str, reference: str) -> str:
        if "!" in reference:
            return reference
        if " " in sheet_name:
            return f"'{sheet_name}'!{reference}"
        return f"{sheet_name}!{reference}"

    def _custom_properties(self, workbook: OpenpyxlWorkbook) -> dict[str, Any]:
        properties: dict[str, Any] = {}
        custom_doc_props = getattr(workbook, "custom_doc_props", None)
        if not custom_doc_props:
            return properties

        for prop in custom_doc_props:
            name = getattr(prop, "name", None)
            value = getattr(prop, "value", None)
            if name:
                properties[str(name)] = value
        return properties

    def _split_keywords(self, keywords: Any) -> list[str]:
        if not keywords:
            return []
        return [item.strip() for item in str(keywords).split(",") if item.strip()]

    def _string_or_none(self, value: Any) -> str | None:
        if value is None:
            return None
        return str(value)

    def _color_to_hex(self, rgb: Any) -> str | None:
        if not rgb:
            return None
        value = str(rgb).replace("0x", "").replace("#", "")
        if len(value) == 8:
            value = value[2:]
        if len(value) != 6:
            return None
        return f"#{value.upper()}"

    def _unique_sheet_name(self, base_name: str, existing: Sequence[str]) -> str:
        existing_lower = {name.lower() for name in existing}
        candidate = base_name[:31] or "Sheet"
        if candidate.lower() not in existing_lower:
            return candidate

        index = 2
        while True:
            suffix = f"_{index}"
            candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
            if candidate.lower() not in existing_lower:
                return candidate
            index += 1

    def _unique_defined_name(self, base_name: str, existing: Iterable[str]) -> str:
        existing_lower = set(existing)
        candidate = base_name or "Name"
        if candidate.lower() not in existing_lower:
            return candidate

        index = 2
        while True:
            candidate = f"{base_name}_{index}"
            if candidate.lower() not in existing_lower:
                return candidate
            index += 1

    def _coordinate_sort_key(self, coordinate: str) -> tuple[int, int]:
        min_col, min_row, _, _ = range_boundaries(f"{coordinate}:{coordinate}")
        return (min_row, min_col)

    def _add_defined_name(self, workbook: OpenpyxlWorkbook, defined_name: Any) -> None:
        defined_names = workbook.defined_names
        if hasattr(defined_names, "add"):
            defined_names.add(defined_name)
            return
        defined_names[defined_name.name] = defined_name


def parse_excel(path: str | Path, **kwargs: Any) -> DocumentContent:
    """Convenience wrapper around :class:`ExcelProcessor.parse`."""

    return ExcelProcessor(**kwargs).parse(path)


def excel_to_dataframe(
    path: str | Path,
    sheet: str | int = 0,
    **kwargs: Any,
) -> Any:
    """Convenience wrapper returning a pandas DataFrame."""

    processor = ExcelProcessor(**kwargs)
    processor.load(path)
    if isinstance(sheet, int):
        sheet = processor.extract_sheets()[sheet]
    return processor.to_dataframe(sheet)
